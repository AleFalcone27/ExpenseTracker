import dearpygui.dearpygui as dpg
from datetime import date
from openpyxl import load_workbook

dpg.create_context() # Creamos contexto para poder acceder a la libreria 

# Archivo en el cual queremos operar
archivo = "gastos1.xlsx"

# Cargamos el archivo 
wb = load_workbook(archivo)

# Agarramos la pestaña activa 
ws = wb.active

# Cuando el archio sea nuevo escribe los headings
a1 = ws["A1"].value
if a1 == None:
    ws["A1"] = "FECHA"
    ws["B1"] = "CATEGORIA"
    ws["C1"] = "PRECIO"
    ws["D1"] = "COMENTARIO"
    

# Categorias para el combo (se pueden añadir nuevas)
categorias = ["Comida","Regalos","Compra Super","Ingresos","Facultad","Pc"]

# Obtenemos la fecha
fecha = date.today()

# Realizamos la carga en el archvio excel de los datos ingresados en nuestro programa
def cargar():
    
    # Obtenemos los valores ingresados por el usuario en los ementos de nuestro programa
    precio_ingresado = dpg.get_value(precio)
    categoria_ingresada = dpg.get_value(categoria)
    comentario_ingresado = dpg.get_value(comentario)
    
    # cargamos en el archivo el nuevo gasto
    ws.append([fecha,categoria_ingresada,int(precio_ingresado),comentario_ingresado])   
    
    # Guardamos los cambios
    wb.save(archivo)
    
    
    
with dpg.window(tag="Primary Window"): # Nombre de nuestra ventana
    dpg.add_text("Vamos a ver si ahora llegas a fin de mes")

    # Creamos los elementos de nuestro programa
    precio = dpg.add_input_text(label="PRECIO",tag=1); # Cramos input del precio
    categoria = dpg.add_combo(categorias, tag=2, label="CATEGORIAS"); # Creamos categoria del gasto
    comentario = dpg.add_input_text(label="COMENTARIO", tag=3) # Creamos input del comentario
    dpg.add_button(label="Save", callback= cargar,) # Al presionar el boton llamamos a la funcion encargada de guardar los datos en la planilla

dpg.create_viewport(title='Expenses Tracer ', width = 600, height = 300) # Damos el nombre y medidas a la ventana
dpg.setup_dearpygui() 
dpg.show_viewport() # Mostramos la ventana grafica
dpg.set_primary_window("Primary Window", True)
dpg.start_dearpygui() # Comienza el bucle de renderizado
dpg.destroy_context() # Destruimos el contexto

